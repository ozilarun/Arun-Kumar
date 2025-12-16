import pdfplumber
import pandas as pd
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ===================================================
# MONTH MAP (FROM YOUR CODE)
# ===================================================

MONTH_MAP = {
    "JAN": "January", "FEB": "February", "MAR": "March",
    "APR": "April", "MAY": "May", "JUN": "June",
    "JUL": "July", "AUG": "August", "SEP": "September",
    "OCT": "October", "NOV": "November", "DEC": "December"
}


def get_month_name(pdf_path):
    name = Path(pdf_path).stem.upper()
    for k, v in MONTH_MAP.items():
        if k in name:
            year = re.search(r"(20\d{2})", name)
            return f"{v} {year.group(1)}" if year else v
    return "Unknown Month"


# ===================================================
# 1️⃣ EXTRACTION — CIMB (YOUR LOGIC)
# ===================================================

def extract_cimb(pdf_path):
    txns = []
    seen = set()

    def to_float(x):
        try:
            return float(str(x).replace(",", "").strip())
        except:
            return 0.0

    def valid_date(x):
        return bool(re.match(r"\d{2}/\d{2}/\d{4}", str(x).strip()))

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table:
                continue

            for row in table[1:]:
                row = (row + [""] * 6)[:6]
                date, desc, ref, wd, dep, bal = row

                if not date or not valid_date(date):
                    continue

                desc = str(desc).replace("\n", " ").strip()

                if any(x in desc.lower() for x in [
                    "no of withdrawal", "no of deposit",
                    "total withdrawal", "total deposit",
                    "end of statement", "baki penutup"
                ]):
                    continue

                debit = to_float(wd)
                credit = to_float(dep)
                balance = to_float(bal)

                if debit == 0 and credit == 0:
                    continue

                key = (date, desc, debit, credit, balance)
                if key in seen:
                    continue
                seen.add(key)

                if ref and ref.strip():
                    desc = f"{desc} Ref: {ref.strip()}"

                txns.append({
                    "date": date.strip(),
                    "description": desc,
                    "debit": debit,
                    "credit": credit,
                    "balance": balance
                })

    return pd.DataFrame(txns, columns=[
        "date", "description", "debit", "credit", "balance"
    ])


# ===================================================
# 2️⃣ OPENING BALANCE (REVERSE FIRST TXN)
# ===================================================

def compute_opening_balance(df):
    first = df.iloc[0]
    return first["balance"] + first["debit"] - first["credit"]


# ===================================================
# 3️⃣ SUMMARY CALCULATION (OD LIMIT FROM USER)
# ===================================================

def compute_cimb_summary(df, od_limit):
    opening = compute_opening_balance(df)
    ending = df.iloc[-1]["balance"]

    total_debit = df["debit"].sum()
    total_credit = df["credit"].sum()

    highest = df["balance"].max()
    lowest = df["balance"].min()
    swing = abs(highest - lowest)

    if od_limit > 0:
        od_util = abs(ending) if ending < 0 else 0
        od_pct = (od_util / od_limit) * 100
    else:
        od_util = 0
        od_pct = 0

    return {
        "Opening Balance": opening,
        "Total Debit": total_debit,
        "Total Credit": total_credit,
        "Ending Balance": ending,
        "Highest Balance": highest,
        "Lowest Balance": lowest,
        "Swing": swing,
        "OD Util (RM)": od_util,
        "OD %": od_pct
    }


# ===================================================
# 4️⃣ EXCEL EXPORT — MATCHES YOUR FORMAT
# ===================================================

def export_cimb_excel(summary_df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Analysis"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    thin = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.append(["CIMB STATEMENT ANALYSIS"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])

    start_row = ws.max_row + 1

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(
        min_row=start_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(summary_df.columns)
    ):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_file)
    return output_file
