import streamlit as st
import tempfile
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# ===============================
# BANK IMPORTS
# ===============================
from bank_rakyat import extract_bank_rakyat
from bank_islam import extract_bank_islam
from cimb import extract_cimb
from maybank import extract_maybank
from rhb import extract_rhb
from ambank import extract_ambank
from agro_bank import extract_agro_bank
from bank_muamalat import extract_bank_muamalat
from public_bank import extract_public_bank
from ocbc import extract_ocbc

BANK_EXTRACTORS = {
    "Bank Rakyat": extract_bank_rakyat,
    "Bank Islam": extract_bank_islam,
    "CIMB": extract_cimb,
    "Maybank": extract_maybank,
    "RHB": extract_rhb,
    "Ambank": extract_ambank,
    "Agrobank": extract_agro_bank,
    "Bank Muamalat": extract_bank_muamalat,
    "Public Bank": extract_public_bank,
    "OCBC": extract_ocbc,
}

# ===============================
# PAGE CONFIG
# ===============================
st.set_page_config(page_title="Bank Statement Analysis", layout="wide")
st.title("🏦 Bank Statement Analysis")

# ===============================
# INPUTS
# ===============================
bank_choice = st.selectbox("Select Bank", list(BANK_EXTRACTORS.keys()))
uploaded_files = st.file_uploader(
    "Upload Bank Statement PDF(s)",
    type=["pdf"],
    accept_multiple_files=True
)

OD_LIMIT = st.number_input("Enter OD Limit (RM)", min_value=0.0, step=1000.0)

if not uploaded_files:
    st.stop()

# ===============================
# EXTRACT PER FILE
# ===============================
extractor = BANK_EXTRACTORS[bank_choice]
monthly_data = {}

for f in uploaded_files:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        tmp.write(f.read())
        path = tmp.name

    df = extractor(path)
    if df is None or df.empty:
        continue

    df["_dt"] = pd.to_datetime(df["date"], dayfirst=True, errors="coerce")
    valid_dates = df["_dt"].dropna()

    if not valid_dates.empty:
        period = valid_dates.iloc[0].to_period("M")
    else:
        m = re.search(
            r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC|"
            r"January|February|March|April|May|June|July|August|"
            r"September|October|November|December)",
            f.name,
            re.IGNORECASE
        )
        if not m:
            raise ValueError(f"Cannot infer month from filename: {f.name}")

        month_str = m.group(1)[:3].title()
        period = pd.to_datetime(f"{month_str} 2023", format="%b %Y").to_period("M")

    label = period.strftime("%b %Y")
    df = df.drop(columns="_dt", errors="ignore").reset_index(drop=True)
    monthly_data[label] = df

# ===============================
# SORT MONTHS
# ===============================
def sort_months(d):
    items = []
    for k, v in d.items():
        items.append((pd.to_datetime(k, format="%b %Y"), k, v))
    items.sort(key=lambda x: x[0])
    return [(k, v) for _, k, v in items]

months = sort_months(monthly_data)

# ===============================
# DISPLAY DATA
# ===============================
st.subheader("📄 Extracted Monthly Data")
for month, df in months:
    with st.expander(month):
        st.dataframe(df, use_container_width=True)

# ===============================
# MONTHLY SUMMARY - FIXED LOGIC
# ===============================
rows = []
previous_ending = None

for idx, (month, df) in enumerate(months):
    first = df.iloc[0]
    last = df.iloc[-1]

    # CORRECTED: Opening balance logic
    # For first month: use first row balance - credit + debit
    # For subsequent months: use previous month's ending balance
    if idx == 0:
        opening = first["balance"] - first["credit"] + first["debit"]
    else:
        opening = previous_ending

    ending = last["balance"]
    
    # Store for next iteration
    previous_ending = ending

    highest = df["balance"].max()
    lowest = df["balance"].min()

    # Total debit and credit for the month
    total_debit = df["debit"].sum()
    total_credit = df["credit"].sum()

    # CORRECTED: OD Utilization (only if balance is negative)
    od_util_rm = abs(ending) if ending < 0 else 0.0
    od_percent = (od_util_rm / OD_LIMIT * 100) if OD_LIMIT > 0 and ending < 0 else 0.0

    rows.append({
        "Month": month,
        "Opening": round(opening, 2),
        "Debit": round(total_debit, 2),
        "Credit": round(total_credit, 2),
        "Ending": round(ending, 2),
        "Highest": round(highest, 2),
        "Lowest": round(lowest, 2),
        "Swing": round(highest - lowest, 2),
        "OD Util (RM)": round(od_util_rm, 2),
        "OD %": round(od_percent, 2)
    })

summary_df = pd.DataFrame(rows)

st.subheader("📅 Summary Table")
st.dataframe(summary_df, use_container_width=True)

# ===============================
# FINANCIAL RATIOS - CORRECTED
# ===============================
total_credit_6m = summary_df["Credit"].sum()
total_debit_6m = summary_df["Debit"].sum()
avg_opening = summary_df["Opening"].mean()
avg_ending = summary_df["Ending"].mean()
highest_period = summary_df["Highest"].max()
lowest_period = summary_df["Lowest"].min()
avg_od_util_rm = summary_df["OD Util (RM)"].mean()
avg_od_percent = summary_df["OD %"].mean()
avg_swing = summary_df["Swing"].mean()

# Calculate % of Swing
percent_of_swing = (avg_swing / OD_LIMIT * 100) if OD_LIMIT > 0 else 0.0

# Count number of months where ending balance exceeds OD limit (is more negative than limit)
num_excesses = int((summary_df["Ending"] < -OD_LIMIT).sum()) if OD_LIMIT > 0 else 0

ratio = {
    "Total Credit (6 Months)": round(total_credit_6m, 2),
    "Total Debit (6 Months)": round(total_debit_6m, 2),
    "Annualized Credit": round(total_credit_6m * 2, 2),
    "Annualized Debit": round(total_debit_6m * 2, 2),
    "Average Opening Balance": round(avg_opening, 2),
    "Average Ending Balance": round(avg_ending, 2),
    "Highest Balance (Period)": round(highest_period, 2),
    "Lowest Balance (Period)": round(lowest_period, 2),
    "Average OD Utilization (RM)": round(avg_od_util_rm, 2),
    "Average % OD Utilization": round(avg_od_percent, 2),
    "Average Monthly Swing (RM)": round(avg_swing, 2),
    "% of Swing": round(percent_of_swing, 2),
    "Returned Cheques": 0,
    "Number of Excesses": num_excesses
}

ratio_df = pd.DataFrame(ratio.items(), columns=["Metric", "Value"])

st.subheader("📊 Financial Ratios")
st.dataframe(ratio_df, use_container_width=True)

# ===============================
# EXCEL EXPORT
# ===============================
wb = Workbook()
ws = wb.active
ws.title = "Summary"

header = PatternFill("solid", fgColor="1F4E78")
font = Font(color="FFFFFF", bold=True)

for r, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=v)
        if r == 1:
            cell.fill = header
            cell.font = font

start = ws.max_row + 3
ws.cell(row=start, column=1, value="Financial Ratios").font = Font(bold=True)

for r, row in enumerate(dataframe_to_rows(ratio_df, index=False, header=True), start + 1):
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)

bio = BytesIO()
wb.save(bio)
bio.seek(0)

st.download_button(
    "⬇ Download Summary + Ratios (Excel)",
    data=bio,
    file_name="Bank_Statement_Analysis.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
