import pdfplumber
import pytesseract
from PIL import Image
import regex as re
import pandas as pd
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ===================================================
# OCR TEMP DIR
# ===================================================

TEMP_DIR = "temp_ocr_images"
os.makedirs(TEMP_DIR, exist_ok=True)


# ===================================================
# HELPERS
# ===================================================

def num(x):
    try:
        x = str(x).replace(",", "").strip()
        if x.endswith("-"):
            return -float(x[:-1])
        if x.endswith("+"):
            return float(x[:-1])
        return float(x)
    except:
        return 0.0


def extract_text(page, page_num):
    text = page.extract_text()
    if text and text.strip():
        return text

    img_path = f"{TEMP_DIR}/page_{page_num}.png"
    page.to_image(resolution=300).save(img_path)
    return pytesseract.image_to_string(Image.open(img_path))


def preprocess_rhb_text(text):
    lines = text.split("\n")
    merged, buffer = [], ""

    for line in lines:
        line = line.strip()
        if re.match(r"^\d{2}-\d{2}-\d{4}", line):
            if buffer:
                merged.append(buffer.strip())
            buffer = line
        else:
            buffer += " " + line

    if buffer:
        merged.append(buffer.strip())

    return "\n".join(merged)


# ===================================================
# TRANSACTION REGEX (YOUR LOGIC)
# ===================================================

txn_pattern = re.compile(
    r"""
    (?P<date>\d{2}-\d{2}-\d{4})\s+
    (?P<body>.*?)\s+
    (?P<dr>[0-9,]*\.\d{2})?\s*(?P<dr_flag>-)?\s*
    (?P<cr>[0-9,]*\.\d{2})?\s+
    (?P<bal>-?[0-9,]*\.\d{2}[+-]?)
    """,
    re.VERBOSE | re.DOTALL
)


# ===================================================
# 1️⃣ EXTRACTION ENTRY (STREAMLIT)
# ===================================================

def extract_rhb(pdf_path):
    txns = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            raw = extract_text(page, page_num)
            processed = preprocess_rhb_text(raw)

            for m in txn_pattern.finditer(processed):
                txns.append({
                    "date": m.group("date"),
                    "description": m.group("body").strip(),
                    "debit": num(m.group("dr")) if m.group("dr") else 0.0,
                    "credit": num(m.group("cr")) if m.group("cr") else 0.0,
                    "balance": num(m.group("bal")),
                })

    return pd.DataFrame(txns, columns=[
        "date", "description", "debit", "credit", "balance"
    ])


# ===================================================
# 2️⃣ OPENING BALANCE
# ===================================================

def compute_opening_balance(df):
    first = df.iloc[0]
    return first["balance"] + first["debit"] - first["credit"]


# ===================================================
# 3️⃣ MONTHLY / PERIOD SUMMARY
# ===================================================

def compute_rhb_summary(df, od_limit):
    opening = compute_opening_balance(df)
    ending = df.iloc[-1]["balance"]

    total_debit = df["debit"].sum()
    total_credit = df["credit"].sum()

    highest = df["balance"].max()
    lowest = df["balance"].min()
    swing = abs(highest - lowest)

    od_util = abs(ending) if ending < 0 else 0
    od_pct = (od_util / od_limit) * 100 if od_limit > 0 else 0

    return {
        "Opening": opening,
        "Debit": total_debit,
        "Credit": total_credit,
        "Ending": ending,
        "Highest": highest,
        "Lowest": lowest,
        "Swing": swing,
        "OD Util (RM)": od_util,
        "OD %": od_pct
    }


# ===================================================
# 4️⃣ RATIOS (FROM YOUR NOTEBOOK)
# ===================================================

def compute_rhb_ratios(summary_df, od_limit):
    ratio = {}

    ratio["Total Credit (6 Months)"] = summary_df["Credit"].sum()
    ratio["Total Debit (6 Months)"] = summary_df["Debit"].sum()
    ratio["Annualized Credit"] = ratio["Total Credit (6 Months)"] * 2
    ratio["Annualized Debit"] = ratio["Total Debit (6 Months)"] * 2
    ratio["Average Opening Balance"] = summary_df["Opening"].mean()
    ratio["Average Ending Balance"] = summary_df["Ending"].mean()
    ratio["Highest Balance (Period)"] = summary_df["Highest"].max()
    ratio["Lowest Balance (Period)"] = summary_df["Lowest"].min()
    ratio["Average OD Util (RM)"] = summary_df["OD Util (RM)"].mean()
    ratio["Average OD %"] = summary_df["OD %"].mean()
    ratio["Average Swing"] = summary_df["Swing"].mean()

    ratio["% of Swing"] = (
        ratio["Average Swing"] / od_limit * 100
        if od_limit > 0 else 0
    )

    return pd.DataFrame(list(ratio.items()), columns=["Metric", "Value"])


# ===================================================
# 5️⃣ EXCEL EXPORT
# ===================================================

def export_rhb_excel(summary_df, ratio_df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Analysis"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["RHB STATEMENT ANALYSIS"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    ws.append([])
    ws.append(["FINANCIAL RATIOS"])
    ws.append([])

    for r in dataframe_to_rows(ratio_df, index=False, header=True):
        ws.append(r)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    wb.save(output_file)
    return output_file
