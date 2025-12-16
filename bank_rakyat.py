import pdfplumber
import pandas as pd
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ===================================================
# 1️⃣ EXTRACTION — BANK RAKYAT (TABULATE SOURCE)
# ===================================================

def extract_bank_rakyat(pdf_path):
    txns = []

    def to_float(x):
        try:
            return float(str(x).replace(",", "").strip())
        except:
            return 0.0

    def valid_date(x):
        return bool(re.match(r"\d{2}/\d{2}/\d{4}", str(x).strip()))

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if not table or len(table) < 2:
                continue

            for row in table[1:]:
                row = [(c or "").strip() for c in row]

                if len(row) < 6:
                    continue

                date, _, desc, debit, credit, balance = row[:6]

                skip_words = [
                    "BAKI PERMULAAN",
                    "BAKI PENUTUP",
                    "JUMLAH",
                    "TOTAL",
                    "BIL",
                    "NO"
                ]

                if any(w in desc.upper() for w in skip_words):
                    continue

                if not valid_date(date):
                    continue

                desc = re.sub(r"\s+", " ", desc).strip()

                txns.append({
                    "date": date,
                    "description": desc,
                    "debit": to_float(debit),
                    "credit": to_float(credit),
                    "balance": to_float(balance),
                })

    return pd.DataFrame(txns, columns=[
        "date", "description", "debit", "credit", "balance"
    ])


# ===================================================
# 2️⃣ OPENING BALANCE (REVERSE FIRST TXN)
# ===================================================

def compute_opening_balance(df):
    first = df.iloc[0]
    return first["balance"] + first["debit"] - first["credit"]


# ===================================================
# 3️⃣ FULL MONTH / PERIOD CALCULATION
# ===================================================

def compute_bank_rakyat_summary(df, od_limit):
    opening = compute_opening_balance(df)
    ending = df.iloc[-1]["balance"]

    total_debit = df["debit"].sum()
    total_credit = df["credit"].sum()

    highest = df["balance"].max()
    lowest = df["balance"].min()
    swing = abs(highest - lowest)

    if od_limit > 0:
        od_util = abs(ending) if ending < 0 else 0
        od_pct = (od_util / od_limit) * 100
    else:
        od_util = 0
        od_pct = 0

    return {
        "Opening Balance": opening,
        "Total Debit": total_debit,
        "Total Credit": total_credit,
        "Ending Balance": ending,
        "Highest Balance": highest,
        "Lowest Balance": lowest,
        "Swing": swing,
        "OD Util (RM)": od_util,
        "OD %": od_pct
    }


# ===================================================
# 4️⃣ EXCEL EXPORT — MATCHES YOUR NOTEBOOK STYLE
# ===================================================

def export_bank_rakyat_excel(summary_df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Analysis"

    # -----------------------
    # Styles
    # -----------------------
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    thin = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # -----------------------
    # Title
    # -----------------------
    ws.append(["BANK RAKYAT STATEMENT ANALYSIS"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])

    start_row = ws.max_row + 1

    # -----------------------
    # Write Summary Table
    # -----------------------
    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    # Header formatting
    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Body formatting
    for row in ws.iter_rows(
        min_row=start_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(summary_df.columns)
    ):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    # -----------------------
    # Auto column width
    # -----------------------
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_file)
    return output_file
