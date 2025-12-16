import pdfplumber
import pandas as pd
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ===================================================
# MONTH MAP
# ===================================================

MONTH_MAP = {
    "01": "January", "02": "February", "03": "March",
    "04": "April",   "05": "May",      "06": "June",
    "07": "July",    "08": "August",   "09": "September",
    "10": "October", "11": "November", "12": "December"
}


# ===================================================
# DETECT MONTH FROM DATAFRAME
# ===================================================

def detect_month_from_df(df):
    if df.empty:
        return "Unknown Month"

    try:
        d = df.iloc[0]["date"]
        dd, mm, yy = d.split("/")
        return f"{MONTH_MAP.get(mm, 'Unknown')} {yy}"
    except:
        return "Unknown Month"


# ===================================================
# 1️⃣ EXTRACTION — BANK ISLAM (CASA + NORMAL)
# ===================================================

def extract_bank_islam(pdf_path):
    txns = []

    def to_float(v):
        try:
            s = str(v).replace("\n", "").replace(",", "")
            return float(s)
        except:
            return 0.0

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:

            table = page.extract_table()
            if not table:
                continue

            is_casa = len(table[0]) >= 11

            for row in table[1:]:

                if not row or all(c in [None, ""] for c in row):
                    continue

                # ======================
                # CASA FORMAT
                # ======================
                if is_casa:
                    try:
                        raw_date = row[1]
                        desc = row[4]
                        debit = row[7]
                        credit = row[8]
                        balance = row[9]

                        if not raw_date:
                            continue

                        date = raw_date.split("\n")[0].strip()
                        if not re.match(r"\d{2}/\d{2}/\d{4}", date):
                            continue

                        txns.append({
                            "date": date,
                            "description": desc.replace("\n", " ").strip(),
                            "debit": to_float(debit),
                            "credit": to_float(credit),
                            "balance": to_float(balance),
                        })
                    except:
                        continue

                # ======================
                # NORMAL FORMAT
                # ======================
                else:
                    try:
                        date = row[0]
                        desc = row[1]
                        debit = row[2]
                        credit = row[3]
                        balance = row[4]

                        if not date or not re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", date):
                            continue

                        txns.append({
                            "date": date.strip(),
                            "description": desc.replace("\n", " ").strip(),
                            "debit": to_float(debit),
                            "credit": to_float(credit),
                            "balance": to_float(balance),
                        })
                    except:
                        continue

    return pd.DataFrame(txns, columns=[
        "date", "description", "debit", "credit", "balance"
    ])


# ===================================================
# 2️⃣ OPENING BALANCE
# ===================================================

def compute_opening_balance(df):
    first = df.iloc[0]
    return first["balance"] + first["debit"] - first["credit"]


# ===================================================
# 3️⃣ SUMMARY CALCULATION (OD LIMIT FROM USER)
# ===================================================

def compute_bank_islam_summary(df, od_limit):
    opening = compute_opening_balance(df)
    ending = df.iloc[-1]["balance"]

    total_debit = df["debit"].sum()
    total_credit = df["credit"].sum()

    highest = df["balance"].max()
    lowest = df["balance"].min()
    swing = abs(highest - lowest)

    if od_limit > 0:
        od_util = abs(ending) if ending < 0 else 0
        od_pct = (od_util / od_limit) * 100
    else:
        od_util = 0
        od_pct = 0

    return {
        "Opening Balance": opening,
        "Total Debit": total_debit,
        "Total Credit": total_credit,
        "Ending Balance": ending,
        "Highest Balance": highest,
        "Lowest Balance": lowest,
        "Swing": swing,
        "OD Util (RM)": od_util,
        "OD %": od_pct
    }


# ===================================================
# 4️⃣ EXCEL EXPORT (MATCHES YOUR FORMAT)
# ===================================================

def export_bank_islam_excel(summary_df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Analysis"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    thin = Side(border_style="thin", color="000000")
    thin_border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws.append(["BANK ISLAM STATEMENT ANALYSIS"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])

    start_row = ws.max_row + 1

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for row in ws.iter_rows(
        min_row=start_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=len(summary_df.columns)
    ):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_file)
    return output_file
