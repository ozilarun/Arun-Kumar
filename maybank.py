import pdfplumber
import pandas as pd
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


# ============================================================
# COMMON HELPERS
# ============================================================

def to_float(x):
    try:
        return float(str(x).replace(",", "").strip())
    except:
        return 0.0


def compute_opening_balance(df):
    first = df.iloc[0]
    return first["balance"] + first["debit"] - first["credit"]


# ============================================================
# MAYBANK CWS EXTRACTOR (FROM YOUR CODE)
# ============================================================

DATE_PATTERN_CWS = re.compile(r"^\d{2}\s+[A-Za-z]{3}\s+\d{4}")
AMOUNT_PATTERN_CWS = re.compile(r'([0-9,]+\.\d{2})\s*([+-])\s*([0-9,]+\.\d{2})')


def extract_maybank_cws(pdf_path):
    txns = []
    current = None
    buffer_desc = []

    with pdfplumber.open(pdf_path) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            text = page.extract_text()
            if not text:
                continue

            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    continue

                if DATE_PATTERN_CWS.match(line):
                    if current:
                        current["description"] = " ".join(buffer_desc).strip()
                        txns.append(current)

                    buffer_desc = []

                    m = AMOUNT_PATTERN_CWS.search(line)
                    if not m:
                        continue

                    amt = to_float(m.group(1))
                    sign = m.group(2)
                    bal = to_float(m.group(3))

                    debit = amt if sign == "-" else 0.0
                    credit = amt if sign == "+" else 0.0

                    current = {
                        "date": line[:11],
                        "description": "",
                        "debit": debit,
                        "credit": credit,
                        "balance": bal,
                    }

                    buffer_desc.append(line[:m.start()].strip())
                else:
                    if current:
                        buffer_desc.append(line)

        if current:
            current["description"] = " ".join(buffer_desc).strip()
            txns.append(current)

    return pd.DataFrame(txns)


# ============================================================
# MAYBANK SME / LARNEY EXTRACTOR (FROM YOUR CODE)
# ============================================================

DATE_PATTERN_SME = re.compile(r"^\d{2}/\d{2}")
AMOUNT_PATTERN_SME = re.compile(r'([0-9,]+\.\d{2})\s*([+-])\s*([0-9,]+\.\d{2})')


def extract_maybank_sme(pdf_path):
    txns = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            lines = text.splitlines()
            current = None
            buffer_desc = []

            for raw_line in lines:
                line = raw_line.strip()
                if not line:
                    continue

                if not DATE_PATTERN_SME.match(line):
                    if buffer_desc:
                        buffer_desc.append(line)
                    continue

                m = AMOUNT_PATTERN_SME.search(line)
                if not m:
                    buffer_desc.append(line)
                    continue

                if current:
                    current["description"] = " ".join(buffer_desc).strip()
                    txns.append(current)

                buffer_desc = [line]

                amt = to_float(m.group(1))
                sign = m.group(2)
                bal = to_float(m.group(3))

                debit = amt if sign == "-" else 0.0
                credit = amt if sign == "+" else 0.0

                current = {
                    "date": line[:5],
                    "description": "",
                    "debit": debit,
                    "credit": credit,
                    "balance": bal,
                }

            if current:
                current["description"] = " ".join(buffer_desc).strip()
                txns.append(current)

    return pd.DataFrame(txns)


# ============================================================
# 1️⃣ STREAMLIT ENTRY POINT (AUTO CWS / SME)
# ============================================================

def extract_maybank(pdf_path):
    df_cws = extract_maybank_cws(pdf_path)

    if not df_cws.empty:
        return df_cws[["date", "description", "debit", "credit", "balance"]]

    df_sme = extract_maybank_sme(pdf_path)

    return df_sme[["date", "description", "debit", "credit", "balance"]]


# ============================================================
# 2️⃣ MONTHLY SUMMARY
# ============================================================

def compute_maybank_summary(df, od_limit):
    opening = compute_opening_balance(df)
    ending = df.iloc[-1]["balance"]

    total_debit = df["debit"].sum()
    total_credit = df["credit"].sum()

    highest = df["balance"].max()
    lowest = df["balance"].min()
    swing = abs(highest - lowest)

    if od_limit > 0:
        od_util = abs(ending) if ending < 0 else 0
        od_pct = (od_util / od_limit) * 100
    else:
        od_util = 0
        od_pct = 0

    return {
        "Opening": opening,
        "Debit": total_debit,
        "Credit": total_credit,
        "Ending": ending,
        "Highest": highest,
        "Lowest": lowest,
        "Swing": swing,
        "OD Util (RM)": od_util,
        "OD %": od_pct
    }


# ============================================================
# 3️⃣ RATIOS (MATCHES YOUR LOGIC)
# ============================================================

def compute_maybank_ratios(summary_df, od_limit):
    ratio = {}

    ratio["Total Credit (6 Months)"] = summary_df["Credit"].sum()
    ratio["Total Debit (6 Months)"] = summary_df["Debit"].sum()
    ratio["Annualized Credit"] = ratio["Total Credit (6 Months)"] * 2
    ratio["Annualized Debit"] = ratio["Total Debit (6 Months)"] * 2
    ratio["Average Opening Balance"] = summary_df["Opening"].mean()
    ratio["Average Ending Balance"] = summary_df["Ending"].mean()
    ratio["Highest Balance (Period)"] = summary_df["Highest"].max()
    ratio["Lowest Balance (Period)"] = summary_df["Lowest"].min()
    ratio["Average OD Utilization (RM)"] = summary_df["OD Util (RM)"].mean()
    ratio["Average % OD Utilization"] = summary_df["OD %"].mean()
    ratio["Average Monthly Swing"] = summary_df["Swing"].mean()

    ratio["% of Swing"] = (
        (ratio["Average Monthly Swing"] / od_limit) * 100
        if od_limit > 0 else 0
    )

    return pd.DataFrame(list(ratio.items()), columns=["Metric", "Value"])


# ============================================================
# 4️⃣ EXCEL EXPORT
# ============================================================

def export_maybank_excel(summary_df, ratio_df, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Analysis"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(["MAYBANK STATEMENT ANALYSIS"])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws.append(r)

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    ws.append([])
    ws.append([])

    ws.append(["FINANCIAL RATIOS"])
    ws["A" + str(ws.max_row)].font = Font(size=14, bold=True)
    ws.append([])

    for r in dataframe_to_rows(ratio_df, index=False, header=True):
        ws.append(r)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    wb.save(output_file)
    return output_file
